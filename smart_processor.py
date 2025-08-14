import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path
import logging
import re
from config import Config


class SmartReportProcessor:
    def __init__(self):
        self.config = Config()
        self.setup_logging()
        self.ensure_directories()

    def setup_logging(self):
        """设置日志"""
        logging.basicConfig(
            level=logging.DEBUG,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('processor.log', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

    def ensure_directories(self):
        """确保必要的目录存在"""
        for dir_path in [self.config.SOURCE_DIR, self.config.TEMPLATE_DIR, self.config.OUTPUT_DIR]:
            Path(dir_path).mkdir(exist_ok=True)

    def debug_dataframe(self, df, title="DataFrame调试信息"):
        """调试DataFrame内容"""
        self.logger.debug(f"\n=== {title} ===")
        self.logger.debug(f"DataFrame形状: {df.shape}")
        self.logger.debug(f"前10行数据:")
        for i in range(min(10, len(df))):
            row_data = df.iloc[i].tolist()
            self.logger.debug(f"第{i}行: {row_data}")
        self.logger.debug("=" * 50)

    def read_source_data(self, file_path):
        """读取原始数据"""
        try:
            df = pd.read_excel(file_path, sheet_name=self.config.SOURCE_SHEET_NAME, header=None)
            self.logger.info(f"成功读取源文件: {file_path}")
            self.debug_dataframe(df, f"原始数据 - {file_path.name}")
            return df
        except Exception as e:
            self.logger.error(f"读取源文件失败 {file_path}: {str(e)}")
            return None

    def extract_test_info(self, df):
        """提取测试信息 - 使用配置的位置"""
        test_info = {}
        pos = self.config.SOURCE_DATA_POSITIONS

        try:
            self.logger.debug(f"开始提取测试信息，使用位置配置: {pos}")

            required_rows = max(pos['item_name_row'], pos['bias1_row'], pos['bias2_row'],
                                pos['bias3_row'], pos['min_limit_row'], pos['max_limit_row'])

            if len(df) <= required_rows:
                self.logger.error(f"DataFrame行数不足，需要至少{required_rows + 1}行，实际只有{len(df)}行")
                return test_info

            item_names = df.iloc[pos['item_name_row'], pos['test_items_start_col']:].tolist()
            bias1_data = df.iloc[pos['bias1_row'], pos['test_items_start_col']:].tolist()
            bias2_data = df.iloc[pos['bias2_row'], pos['test_items_start_col']:].tolist()
            bias3_data = df.iloc[pos['bias3_row'], pos['test_items_start_col']:].tolist()
            min_limits = df.iloc[pos['min_limit_row'], pos['test_items_start_col']:].tolist()
            max_limits = df.iloc[pos['max_limit_row'], pos['test_items_start_col']:].tolist()

            self.logger.debug(f"测试项目名称: {item_names[:5]}...")
            self.logger.debug(f"Bias1数据: {bias1_data[:5]}...")
            self.logger.debug(f"最小限值: {min_limits[:5]}...")
            self.logger.debug(f"最大限值: {max_limits[:5]}...")

            for i, item_name in enumerate(item_names):
                if pd.notna(item_name) and str(item_name).strip():
                    clean_name = str(item_name).strip()
                    test_info[clean_name] = {
                        'bias1': bias1_data[i] if i < len(bias1_data) else '',
                        'bias2': bias2_data[i] if i < len(bias2_data) else '',
                        'bias3': bias3_data[i] if i < len(bias3_data) else '',
                        'min_limit': min_limits[i] if i < len(min_limits) else None,
                        'max_limit': max_limits[i] if i < len(max_limits) else None,
                        'column_index': pos['test_items_start_col'] + i
                    }
                    self.logger.debug(f"添加测试项: {clean_name} -> 列{pos['test_items_start_col'] + i}")

            self.logger.info(f"提取到 {len(test_info)} 个测试项目")
            self.logger.debug(f"测试项目列表: {list(test_info.keys())}")

        except Exception as e:
            self.logger.error(f"提取测试信息失败: {str(e)}")
            self.logger.exception("详细错误:")

        return test_info

    def extract_test_data(self, df):
        """提取测试数据 - 支持P或F前缀"""
        test_data = []
        pos = self.config.SOURCE_DATA_POSITIONS
        recognition = self.config.DATA_RECOGNITION

        try:
            self.logger.debug(f"开始提取测试数据，从第{pos['data_start_row']}行开始")

            # 🆕 获取支持的前缀，支持P或F
            sample_prefix = recognition['sample_prefix']
            if isinstance(sample_prefix, str):
                # 如果是单个前缀，转换为列表以支持多前缀检查
                supported_prefixes = [sample_prefix]
            else:
                # 如果已经是列表，直接使用
                supported_prefixes = sample_prefix

            # 🆕 添加F前缀支持
            if 'P' in supported_prefixes and 'F' not in supported_prefixes:
                supported_prefixes.append('F')
            elif 'F' in supported_prefixes and 'P' not in supported_prefixes:
                supported_prefixes.append('P')
            elif sample_prefix == 'P':
                supported_prefixes = ['P', 'F']

            self.logger.debug(f"支持的样品前缀: {supported_prefixes}")

            start_row = pos['data_start_row']
            max_rows = start_row + recognition['max_data_rows']

            for idx in range(start_row, min(len(df), max_rows)):
                row_data = df.iloc[idx].tolist()

                if len(row_data) > pos['sample_id_col']:
                    sample_id = row_data[pos['sample_id_col']]

                    self.logger.debug(f"第{idx}行，样品ID: {sample_id}")

                    # 🆕 检查是否匹配任何支持的前缀
                    if pd.notna(sample_id):
                        sample_id_str = str(sample_id).strip()
                        is_valid_sample = False

                        for prefix in supported_prefixes:
                            if sample_id_str.startswith(prefix):
                                is_valid_sample = True
                                self.logger.debug(f"样品 {sample_id} 匹配前缀 {prefix}")
                                break

                        if is_valid_sample:
                            test_data.append(row_data)
                            self.logger.debug(f"添加测试数据行: {sample_id}")

                    elif recognition['auto_detect_data_end'] and not pd.notna(sample_id):
                        if recognition['skip_empty_rows']:
                            continue
                        else:
                            self.logger.debug(f"遇到空行，停止数据提取")
                            break

            self.logger.info(f"提取到 {len(test_data)} 行测试数据")

            for i, row in enumerate(test_data[:3]):
                self.logger.debug(f"测试数据第{i + 1}行: {row[:10]}...")

        except Exception as e:
            self.logger.error(f"提取测试数据失败: {str(e)}")
            self.logger.exception("详细错误:")

        return test_data

    def map_to_template_items(self, test_info):
        """将原始测试项映射到模板测试项"""
        template_data = {}
        processing = self.config.DATA_PROCESSING

        self.logger.debug(f"开始映射测试项，映射规则: {self.config.TEST_ITEMS_MAPPING}")

        for template_item, source_items in self.config.TEST_ITEMS_MAPPING.items():
            template_data[template_item] = {
                'conditions': [],
                'min_limits': [],
                'max_limits': [],
                'source_columns': []
            }

            self.logger.debug(f"处理模板项: {template_item}")

            for source_item in source_items:
                if source_item in test_info:
                    info = test_info[source_item]

                    # 处理测试条件，支持分行显示
                    condition_parts = []
                    for bias_key in ['bias1', 'bias2', 'bias3']:
                        bias_value = info.get(bias_key, '')
                        if bias_value and str(bias_value).strip() and str(bias_value).strip() != 'nan':
                            condition_parts.append(str(bias_value).strip())

                    # 如果启用分行显示，每个条件单独存储
                    if processing['conditions_multiline']:
                        template_data[template_item]['conditions'].extend(condition_parts)
                    else:
                        condition_text = processing['combine_conditions_separator'].join(condition_parts)
                        template_data[template_item]['conditions'].append(condition_text)

                    template_data[template_item]['min_limits'].append(info['min_limit'])
                    template_data[template_item]['max_limits'].append(info['max_limit'])
                    template_data[template_item]['source_columns'].append(info['column_index'])

                    self.logger.debug(f"  找到源项: {source_item} -> 列{info['column_index']}")
                else:
                    self.logger.warning(f"  未找到源项: {source_item}")

        for item, data in template_data.items():
            self.logger.debug(f"模板项 {item}: 源列{data['source_columns']}, 条件数{len(data['conditions'])}")

        return template_data

    def is_valid_value(self, value):
        """检查值是否有效（包括数值0和"Over"）"""
        if pd.isna(value):
            return False

        value_str = str(value).strip()
        if not value_str or value_str.lower() in ['nan', 'n/a', '']:
            return False

        # 🆕 新增：处理"Over"值
        if value_str.upper() == "OVER":
            return True

        # 数值0是有效值
        try:
            float_val = float(
                value_str.replace('V', '').replace('A', '').replace('R', '').replace('mV', '').replace('uA',
                                                                                                       '').replace('nA',
                                                                                                                   '').replace(
                    'mR', '').replace('ohm', '').replace('Ω', ''))
            return True
        except:
            return len(value_str) > 0

    def convert_to_numeric(self, value_str):
        """将字符串值转换为数值形式，保留0值和"Over"值"""
        if not self.is_valid_value(value_str):
            return None

        value_str = str(value_str).strip()

        # 🆕 新增：直接返回"Over"值
        if value_str.upper() == "OVER":
            return "Over"

        try:
            original_value = value_str
            clean_value = value_str

            # 移除单位符号但保持数值不变
            for unit_type, units in self.config.VALUE_PROCESSING['unit_patterns'].items():
                for unit in units:
                    clean_value = clean_value.replace(unit, '')

            # 转换为数值
            numeric_value = float(clean_value.strip())

            # 根据配置决定精度
            if self.config.DATA_PROCESSING['convert_to_numeric']:
                precision = self.config.VALUE_PROCESSING['decimal_places']
                if abs(numeric_value) < self.config.VALUE_PROCESSING[
                    'scientific_notation_threshold'] and numeric_value != 0:
                    return f"{numeric_value:.{precision}e}"
                else:
                    return round(numeric_value, precision)
            else:
                return original_value

        except (ValueError, TypeError):
            self.logger.debug(f"无法转换为数值: {value_str}")
            return value_str if not self.config.VALUE_PROCESSING['force_numeric_output'] else None

    def clean_numeric_value(self, value_str):
        """清理数值字符串，用于限值比较，"Over"值返回特殊标记"""
        if not self.is_valid_value(value_str):
            return None

        value_str = str(value_str).strip()

        # 🆕 对于Over值，返回无穷大用于比较
        if self.is_over_value(value_str):
            return float('inf')  # 返回无穷大，表示超出所有限值

        try:
            # 移除单位符号但不进行单位转换
            clean_value = value_str
            for unit_type, units in self.config.VALUE_PROCESSING['unit_patterns'].items():
                for unit in units:
                    clean_value = clean_value.replace(unit, '')

            return float(clean_value.strip())
        except (ValueError, TypeError):
            return None

    def count_abnormal_data(self, test_data, template_data):
        """统计异常数据行数（同一行多个异常只计算一次，包括Over值）- 支持P或F前缀，绝对值比较"""
        try:
            if not self.config.ABNORMAL_STATISTICS.get('enable_counting', True):
                return 0

            abnormal_count = 0
            sample_id_col_pos = self.config.SOURCE_DATA_POSITIONS['sample_id_col']
            recognition = self.config.DATA_RECOGNITION
            
            # 获取支持的前缀，支持P或F
            sample_prefix = recognition['sample_prefix']
            if isinstance(sample_prefix, str):
                supported_prefixes = [sample_prefix]
            else:
                supported_prefixes = sample_prefix
            
            # 添加F前缀支持
            if 'P' in supported_prefixes and 'F' not in supported_prefixes:
                supported_prefixes.append('F')
            elif 'F' in supported_prefixes and 'P' not in supported_prefixes:
                supported_prefixes.append('P')
            elif sample_prefix == 'P':
                supported_prefixes = ['P', 'F']

            # 详细统计信息
            abnormal_samples = []
            over_count = 0
            range_abnormal_count = 0

            self.logger.debug(f"开始统计异常数据，数据行数: {len(test_data)}")
            self.logger.debug(f"支持的样品前缀: {supported_prefixes}")
            self.logger.debug("🆕 使用绝对值比较模式")

            for test_row in test_data:
                if len(test_row) <= sample_id_col_pos:
                    continue

                sample_id = str(test_row[sample_id_col_pos])
                
                # 检查是否匹配任何支持的前缀
                is_valid_sample = False
                for prefix in supported_prefixes:
                    if sample_id.strip().startswith(prefix):
                        is_valid_sample = True
                        break
                
                if not is_valid_sample:
                    continue

                # 检查这一行是否有任何异常
                has_abnormal = False
                sample_abnormal_items = []

                for item_name, data in template_data.items():
                    for source_col in data['source_columns']:
                        if source_col < len(test_row):
                            val = test_row[source_col]

                            if self.is_valid_value(val):
                                # 优先检查Over值
                                if self.is_over_value(val):
                                    has_abnormal = True
                                    over_count += 1
                                    sample_abnormal_items.append(f"{item_name}={val}(Over)")
                                    self.logger.debug(f"样品 {sample_id} 发现Over值: {item_name}={val}")
                                    break

                                # 🆕 检查数值范围异常（使用绝对值比较）
                                clean_val = self.clean_numeric_value(val)
                                if clean_val is not None:
                                    if self.is_value_abnormal(clean_val, data):
                                        has_abnormal = True
                                        range_abnormal_count += 1
                                        # 🆕 在日志中显示绝对值比较信息
                                        abs_val = abs(clean_val) if clean_val != float('inf') else clean_val
                                        sample_abnormal_items.append(f"{item_name}={val}(范围,绝对值:{abs_val})")
                                        self.logger.debug(f"样品 {sample_id} 发现范围异常: {item_name}={val} (绝对值:{abs_val})")
                                        break

                    if has_abnormal:
                        break

                if has_abnormal:
                    abnormal_count += 1
                    abnormal_samples.append({
                        'sample_id': sample_id,
                        'abnormal_items': sample_abnormal_items
                    })
                    self.logger.debug(f"发现异常行: {sample_id}")

            # 详细统计日志
            self.logger.info(f"异常统计完成 - 总异常数量: {abnormal_count}")
            self.logger.info(f"统计详情 - Over值: {over_count}, 范围异常(绝对值比较): {range_abnormal_count}")
            self.logger.info(f"异常样品数量: {len(abnormal_samples)}")

            if self.config.LOGGING.get('log_statistics', True):
                for abnormal_sample in abnormal_samples:
                    self.logger.debug(
                        f"异常样品 {abnormal_sample['sample_id']}: {', '.join(abnormal_sample['abnormal_items'])}")

            return abnormal_count

        except Exception as e:
            self.logger.error(f"统计异常数据失败: {str(e)}")
            return 0

    def is_over_value(self, value):
        """检查是否为Over值 - 使用配置的模式"""
        try:
            if isinstance(value, str):
                value_str = str(value).strip()

                # 🆕 使用配置中的Over值模式
                over_patterns = self.config.DATA_RECOGNITION.get('over_value_patterns',
                                                                 ['OVER', 'Over', 'over'])

                for pattern in over_patterns:
                    if pattern in value_str:
                        return True

                # 检查是否以>开头的数值
                if value_str.startswith('>'):
                    return True

            return False

        except Exception as e:
            self.logger.warning(f"检查Over值时出错: {str(e)}")
            return False

    def is_value_abnormal(self, numeric_value, limit_data):
        """判断单个数值是否异常 - 只比较绝对值大小"""
        try:
            # 🆕 取绝对值进行比较
            abs_value = abs(numeric_value) if numeric_value != float('inf') else numeric_value
            
            # 检查最小限值
            for min_limit in limit_data['min_limits']:
                if min_limit is not None:
                    min_val = self.clean_numeric_value(min_limit)
                    if min_val is not None:
                        # 🆕 取绝对值比较
                        abs_min_val = abs(min_val)
                        if abs_value < abs_min_val:
                            self.logger.debug(f"绝对值 {abs_value} 低于最小限值绝对值 {abs_min_val}")
                            return True

            # 检查最大限值
            for max_limit in limit_data['max_limits']:
                if max_limit is not None:
                    max_val = self.clean_numeric_value(max_limit)
                    if max_val is not None:
                        # 🆕 取绝对值比较
                        abs_max_val = abs(max_val)
                        if abs_value > abs_max_val:
                            self.logger.debug(f"绝对值 {abs_value} 超过最大限值绝对值 {abs_max_val}")
                            return True

            return False
        except Exception as e:
            self.logger.warning(f"判断异常值失败: {str(e)}")
            return False

    def filter_group_test_data(self, test_data, start_sample, end_sample):
        """筛选指定数据组的测试数据 - 支持P或F前缀"""
        try:
            filtered_data = []
            recognition = self.config.DATA_RECOGNITION
            sample_id_col_pos = self.config.SOURCE_DATA_POSITIONS['sample_id_col']

            # 🆕 获取支持的前缀，支持P或F
            sample_prefix = recognition['sample_prefix']
            if isinstance(sample_prefix, str):
                supported_prefixes = [sample_prefix]
            else:
                supported_prefixes = sample_prefix

            # 🆕 添加F前缀支持
            if 'P' in supported_prefixes and 'F' not in supported_prefixes:
                supported_prefixes.append('F')
            elif 'F' in supported_prefixes and 'P' not in supported_prefixes:
                supported_prefixes.append('P')
            elif sample_prefix == 'P':
                supported_prefixes = ['P', 'F']

            # 使用主要前缀进行范围显示
            primary_prefix = supported_prefixes[0] if supported_prefixes else 'P'

            self.logger.debug(f"开始筛选数据组 {primary_prefix}{start_sample}-{primary_prefix}{end_sample}")
            self.logger.debug(f"支持的前缀: {supported_prefixes}")

            for test_row in test_data:
                if len(test_row) > sample_id_col_pos:
                    sample_id = str(test_row[sample_id_col_pos])

                    # 🆕 检查是否匹配任何支持的前缀
                    matched_prefix = None
                    sample_num = None

                    for prefix in supported_prefixes:
                        if sample_id.startswith(prefix):
                            matched_prefix = prefix
                            try:
                                # 提取样品编号
                                sample_num_str = sample_id[len(prefix):]
                                sample_num = int(sample_num_str)
                                break
                            except ValueError:
                                self.logger.warning(f"无法解析样品编号: {sample_id}")
                                continue

                    if matched_prefix and sample_num is not None:
                        # 🆕 检查是否在当前数据组范围内
                        if start_sample <= sample_num <= end_sample:
                            filtered_data.append(test_row)
                            self.logger.debug(f"样品 {sample_id} (前缀:{matched_prefix}) 包含在数据组范围内")
                        else:
                            self.logger.debug(f"样品 {sample_id} 不在数据组范围内，跳过")
                    else:
                        self.logger.debug(f"样品 {sample_id} 不符合前缀规则，跳过")

            self.logger.info(
                f"数据组 {primary_prefix}{start_sample}-{primary_prefix}{end_sample} 筛选完成，包含 {len(filtered_data)} 行数据")

            return filtered_data

        except Exception as e:
            self.logger.error(f"筛选数据组数据失败: {str(e)}")
            return []

    def write_to_template(self, template_path, output_path, template_data, test_data):
        """写入模板并生成报告"""
        try:
            self.logger.info(f"开始写入模板: {template_path}")
            self.logger.info(f"总测试数据行数: {len(test_data)}")

            workbook = openpyxl.load_workbook(template_path)
            self.logger.debug(f"模板工作表: {workbook.sheetnames}")

            # 🆕 移除全局异常统计，改为分组统计
            # abnormal_count = self.count_abnormal_data(test_data, template_data)  # 删除这行

            # 🆕 为每个数据组分别统计异常数量
            for group_name, group_config in self.config.DATA_GROUPS.items():
                sheet_index = group_config['target_sheet']
                if sheet_index < len(workbook.worksheets):
                    sheet = workbook.worksheets[sheet_index]
                    self.logger.info(f"处理数据组: {group_name} -> 工作表{sheet_index}({sheet.title})")

                    # 🆕 获取当前数据组的数据范围
                    start_sample, end_sample = group_config['range']
                    self.logger.info(f"数据组范围: P{start_sample}-P{end_sample}")

                    # 🆕 筛选当前数据组的测试数据
                    group_test_data = self.filter_group_test_data(test_data, start_sample, end_sample)
                    self.logger.info(f"数据组 {group_name} 筛选出 {len(group_test_data)} 行数据")

                    # 🆕 统计当前数据组的异常数量
                    group_abnormal_count = self.count_abnormal_data(group_test_data, template_data)

                    # 写入数据组数据
                    self.write_group_data(sheet, template_data, test_data, group_config)

                    # 🆕 写入当前数据组的异常统计
                    self.write_abnormal_count(sheet, group_abnormal_count, sheet_index)

                    self.logger.info(f"数据组 {group_name} 处理完成")
                    self.logger.info(f"  - 样品范围: P{start_sample}-P{end_sample}")
                    self.logger.info(f"  - 数据行数: {len(group_test_data)}")
                    self.logger.info(f"  - 异常数量: {group_abnormal_count}")
                    self.logger.info("-" * 50)
                else:
                    self.logger.error(f"工作表索引{sheet_index}超出范围，总共{len(workbook.worksheets)}个工作表")

            workbook.save(output_path)
            self.logger.info(f"成功生成报告: {output_path}")

        except Exception as e:
            self.logger.error(f"写入模板失败: {str(e)}")
            self.logger.exception("详细错误信息:")

    def write_abnormal_count(self, sheet, abnormal_count, sheet_index=0):
        """写入异常统计到指定位置（数值形式，不改变格式）"""
        try:
            # 检查是否启用异常统计
            if not self.config.ABNORMAL_STATISTICS.get('enable_counting', True):
                self.logger.debug("异常统计功能已禁用，跳过写入")
                return

            if not self.config.ABNORMAL_STATISTICS.get('write_to_template', True):
                self.logger.debug("异常统计写入模板功能已禁用，跳过写入")
                return

            # 获取对应工作表的位置配置
            positions = self.config.ABNORMAL_STATISTICS.get('positions', {})
            sheet_key = f"sheet_{sheet_index}"

            # 如果没有找到对应工作表的配置，使用默认配置
            if sheet_key not in positions:
                self.logger.warning(f"未找到工作表{sheet_index}的异常统计位置配置，使用默认位置")
                position_config = {
                    "row": 1,
                    "col": 1,
                    "write_as_number": True
                }
            else:
                position_config = positions[sheet_key]

            # 获取位置信息
            abnormal_row = position_config.get('row', 1)
            abnormal_col = position_config.get('col', 1)
            write_as_number = position_config.get('write_as_number', True)

            # 🆕 根据配置决定写入格式
            if write_as_number:
                # 🆕 直接写入数值，不改变任何格式
                cell_value = abnormal_count
                display_info = f"数值: {abnormal_count}"
            else:
                # 写入格式化文本（保留原有功能）
                format_template = position_config.get('format', "异常数量: {count}")
                cell_value = format_template.format(count=abnormal_count)
                display_info = f"文本: {cell_value}"

            # 🆕 只写入值，不修改任何格式
            sheet.cell(row=abnormal_row, column=abnormal_col, value=cell_value)

            # 🆕 移除所有格式设置代码，保持原有格式不变
            # 不再设置字体、颜色、对齐等格式

            self.logger.info(f"异常统计已写入工作表{sheet_index}: 行{abnormal_row}, 列{abnormal_col}, {display_info}")

            # 如果启用了详细日志
            if self.config.LOGGING.get('log_statistics', True):
                self.logger.debug(
                    f"异常统计详情 - 工作表: {sheet.title}, 位置: ({abnormal_row}, {abnormal_col}), 数量: {abnormal_count}, 格式: {'数值' if write_as_number else '文本'}")

        except Exception as e:
            self.logger.error(f"写入异常统计失败: {str(e)}")
            self.logger.exception("详细错误信息:")

    def write_group_data(self, sheet, template_data, test_data, group_config):
        """写入分组数据到指定表格，支持P或F前缀"""
        pos = self.config.TEMPLATE_POSITIONS
        processing = self.config.DATA_PROCESSING
        recognition = self.config.DATA_RECOGNITION
        col_offset = pos['test_items_start_col']

        # 🆕 获取支持的前缀，支持P或F
        sample_prefix = recognition['sample_prefix']
        if isinstance(sample_prefix, str):
            supported_prefixes = [sample_prefix]
        else:
            supported_prefixes = sample_prefix

        # 🆕 添加F前缀支持
        if 'P' in supported_prefixes and 'F' not in supported_prefixes:
            supported_prefixes.append('F')
        elif 'F' in supported_prefixes and 'P' not in supported_prefixes:
            supported_prefixes.append('P')
        elif sample_prefix == 'P':
            supported_prefixes = ['P', 'F']

        self.logger.debug(f"写入数据到工作表: {sheet.title}")
        self.logger.debug(f"模板位置配置: {pos}")
        self.logger.debug(f"数据组配置: {group_config}")
        self.logger.debug(f"支持的前缀: {supported_prefixes}")

        # 写入测试项目名称
        for i, item_name in enumerate(self.config.TEST_ITEMS_MAPPING.keys()):
            cell = sheet.cell(row=pos['test_items_row'], column=col_offset + i, value=item_name)
            self.logger.debug(f"写入测试项目: 行{pos['test_items_row']}, 列{col_offset + i}, 值: {item_name}")

        # 写入测试条件，支持分行显示
        for i, (item_name, data) in enumerate(template_data.items()):
            col = col_offset + i

            # 测试条件 - 支持分行显示
            if data['conditions']:
                if processing['conditions_multiline'] and len(data['conditions']) > 1:
                    # 分行显示测试条件
                    for j, condition in enumerate(data['conditions'][:pos['test_conditions_max_rows']]):
                        if condition:
                            condition_row = pos['test_conditions_row'] + j
                            sheet.cell(row=condition_row, column=col, value=condition)
                            self.logger.debug(f"写入测试条件: 行{condition_row}, 列{col}, 值: {condition}")
                else:
                    # 单行显示所有条件
                    conditions_text = processing['combine_conditions_separator'].join(
                        [cond for cond in data['conditions'] if cond]
                    )
                    if conditions_text:
                        sheet.cell(row=pos['test_conditions_row'], column=col, value=conditions_text)
                        self.logger.debug(
                            f"写入测试条件: 行{pos['test_conditions_row']}, 列{col}, 值: {conditions_text}")

            # 规格限值
            if data['min_limits']:
                min_vals = [str(x) for x in data['min_limits'] if x is not None and str(x).strip() != 'nan']
                if min_vals:
                    min_text = processing['combine_values_separator'].join(min_vals)
                    sheet.cell(row=pos['min_limit_row'], column=col, value=min_text)
                    self.logger.debug(f"写入最小限值: 行{pos['min_limit_row']}, 列{col}, 值: {min_text}")

            if data['max_limits']:
                max_vals = [str(x) for x in data['max_limits'] if x is not None and str(x).strip() != 'nan']
                if max_vals:
                    max_text = processing['combine_values_separator'].join(max_vals)
                    sheet.cell(row=pos['max_limit_row'], column=col, value=max_text)
                    self.logger.debug(f"写入最大限值: 行{pos['max_limit_row']}, 列{col}, 值: {max_text}")

        # 写入测试数据，支持P或F前缀
        start_idx, end_idx = group_config['range']
        data_row = pos['data_start_row']
        sample_id_col_pos = self.config.SOURCE_DATA_POSITIONS['sample_id_col']

        self.logger.debug(f"开始写入测试数据，范围: {start_idx}-{end_idx}")

        written_count = 0
        for test_row in test_data:
            if len(test_row) > sample_id_col_pos:
                sample_id = str(test_row[sample_id_col_pos])

                # 🆕 检查是否匹配任何支持的前缀并提取编号
                matched_prefix = None
                p_number = None

                for prefix in supported_prefixes:
                    if sample_id.startswith(prefix):
                        matched_prefix = prefix
                        p_num = sample_id.replace(prefix, '')
                        try:
                            p_number = int(p_num)
                            break
                        except ValueError:
                            self.logger.warning(f"无法解析样品编号: {sample_id}")
                            continue

                if matched_prefix and p_number is not None:
                    if start_idx <= p_number <= end_idx:
                        row_num = p_number - start_idx + 1
                        sheet.cell(row=data_row, column=pos['sample_id_col'], value=row_num)
                        self.logger.debug(f"写入样品{sample_id} (前缀:{matched_prefix}): 行{data_row}, 序号{row_num}")

                        # 写入各测试项的数据
                        for i, (item_name, data) in enumerate(template_data.items()):
                            col = col_offset + i
                            values = []
                            numeric_values = []

                            for source_col in data['source_columns']:
                                if source_col < len(test_row):
                                    val = test_row[source_col]

                                    # 关键修复：使用新的有效性检查，包括0值
                                    if self.is_valid_value(val):
                                        # 转换为数值形式
                                        numeric_val = self.convert_to_numeric(val)
                                        if numeric_val is not None:
                                            values.append(str(numeric_val))
                                            # 用于限值比较的清理数值
                                            clean_val = self.clean_numeric_value(val)
                                            if clean_val is not None:
                                                numeric_values.append(clean_val)

                                        self.logger.debug(f"    原始值: {val}, 转换后: {numeric_val}")

                            # 写入单元格值
                            if values:
                                cell_value = processing['combine_values_separator'].join(values)
                            else:
                                cell_value = processing['empty_value_placeholder']

                            cell = sheet.cell(row=data_row, column=col, value=cell_value)

                            # 如果只有一个数值且启用了数值转换，直接写入数值而不是字符串
                            if (len(values) == 1 and processing['convert_to_numeric']):
                                try:
                                    # 修复：确保0值也能正确写入
                                    if values[0] != "Over":  # 🆕 Over值保持为文本
                                        numeric_cell_value = float(values[0])
                                        cell.value = numeric_cell_value
                                        self.logger.debug(
                                            f"  写入数值: 行{data_row}, 列{col}, 项目{item_name}, 数值: {numeric_cell_value}")
                                    else:
                                        self.logger.debug(
                                            f"  写入Over值: 行{data_row}, 列{col}, 项目{item_name}, 文本: {cell_value}")
                                except:
                                    self.logger.debug(
                                        f"  写入文本: 行{data_row}, 列{col}, 项目{item_name}, 文本: {cell_value}")
                            else:
                                self.logger.debug(
                                    f"  写入数据: 行{data_row}, 列{col}, 项目{item_name}, 值: {cell_value}")

                            # 检查是否超出限值并高亮
                            self.check_and_highlight(cell, numeric_values, data)

                        data_row += 1
                        written_count += 1

        self.logger.info(f"数据组 {group_config.get('description', '')} 写入完成，共写入 {written_count} 行数据")

    def check_and_highlight(self, cell, numeric_values, limit_data):
        """检查数值是否超限并高亮显示（包括"Over"值）- 只比较绝对值大小"""
        if not numeric_values:
            return

        try:
            highlight_fill = PatternFill(
                start_color=self.config.HIGHLIGHT_COLOR,
                end_color=self.config.HIGHLIGHT_COLOR,
                fill_type="solid"
            )

            should_highlight = False

            for i, val in enumerate(numeric_values):
                # 🆕 修改：对于"Over"值（无穷大），直接标记为异常
                if val == float('inf'):
                    should_highlight = True
                    self.logger.debug(f"Over值被标记为异常")
                    continue

                # 🆕 取绝对值进行比较
                abs_val = abs(val)

                # 检查最小限值
                if (i < len(limit_data['min_limits']) and
                        limit_data['min_limits'][i] is not None):

                    min_limit = self.clean_numeric_value(limit_data['min_limits'][i])
                    if min_limit is not None:
                        # 🆕 取绝对值比较
                        abs_min_limit = abs(min_limit)
                        if abs_val < abs_min_limit:
                            should_highlight = True
                            self.logger.debug(f"绝对值 {abs_val} 低于最小限值绝对值 {abs_min_limit} (原值: {val}, 原限值: {min_limit})")

                # 检查最大限值
                if (i < len(limit_data['max_limits']) and
                        limit_data['max_limits'][i] is not None):

                    max_limit = self.clean_numeric_value(limit_data['max_limits'][i])
                    if max_limit is not None:
                        # 🆕 取绝对值比较
                        abs_max_limit = abs(max_limit)
                        if abs_val > abs_max_limit:
                            should_highlight = True
                            self.logger.debug(f"绝对值 {abs_val} 超过最大限值绝对值 {abs_max_limit} (原值: {val}, 原限值: {max_limit})")

            if should_highlight:
                cell.fill = highlight_fill

        except Exception as e:
            self.logger.warning(f"高亮检查失败: {str(e)}")

    def process_all_reports(self):
        """处理所有报告"""
        source_dir = Path(self.config.SOURCE_DIR)
        template_path = Path(self.config.TEMPLATE_DIR) / self.config.TEMPLATE_FILE
        output_dir = Path(self.config.OUTPUT_DIR)

        self.logger.info(f"源文件目录: {source_dir}")
        self.logger.info(f"模板文件: {template_path}")
        self.logger.info(f"输出目录: {output_dir}")

        if not template_path.exists():
            self.logger.error(f"模板文件不存在: {template_path}")
            return

        excel_files = list(source_dir.glob("*.xlsx")) + list(source_dir.glob("*.xls"))

        if not excel_files:
            self.logger.warning(f"在 {source_dir} 中未找到Excel源文件")
            return

        self.logger.info(f"找到 {len(excel_files)} 个Excel文件: {[f.name for f in excel_files]}")

        processed_count = 0
        error_count = 0

        for file_path in excel_files:
            try:
                self.logger.info(f"开始处理: {file_path.name}")

                df = self.read_source_data(file_path)
                if df is None:
                    error_count += 1
                    continue

                test_info = self.extract_test_info(df)
                test_data = self.extract_test_data(df)
                template_data = self.map_to_template_items(test_info)

                output_file = output_dir / f"processed_{file_path.stem}.xlsx"

                self.write_to_template(template_path, output_file, template_data, test_data)
                processed_count += 1

            except Exception as e:
                error_count += 1
                self.logger.error(f"处理文件 {file_path.name} 时发生错误: {str(e)}")
                self.logger.exception("详细错误信息:")

                if not self.config.ERROR_HANDLING['continue_on_error']:
                    break

        self.logger.info(f"处理完成！成功: {processed_count}, 失败: {error_count}")


def main():
    """主函数"""
    processor = SmartReportProcessor()
    processor.process_all_reports()


if __name__ == "__main__":
    main()
